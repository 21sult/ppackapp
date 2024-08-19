import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from streamlit_gsheets import GSheetsConnection

from sklearn.model_selection import train_test_split
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.preprocessing import LabelEncoder


# Setup
## Configurar web page
st.set_page_config(
    page_title = 'Premier Pack - Dashboard',
    page_icon = ':bar_chart:',
    layout = 'wide'
)


## Conectar ao Google Sheets
ttl = 60 # time (seconds) it takes for data to be cleared from cache
conn = st.connection('gsheets', type=GSheetsConnection)
df = conn.read(ttl=10)
#st.write(df)

# Funções
def df_to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Banco de Dados')
    output.seek(0)
    
    # Load workbook and worksheet
    wb = load_workbook(output)
    ws = wb['Banco de Dados']
    
    # Apply styles to header
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    # Save workbook back
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

## Converter data para datetime
df['DATA'] = pd.to_datetime(df['DATA'], format='%Y-%m-%d')

## Título
st.title(':bar_chart: Dashboard Premier Pack')

## Tabs
tabs = st.tabs(['Estatísticas', 'Maiores Faturamentos', 'Tabela', 'Baixar Folha', 'Recomendações'])

with tabs[0]:
    # Sidebar
    ## Definir filtros
    st.sidebar.header('Filtros:')

    if 'show_more_filters' not in st.session_state:
        st.session_state.show_more_filters = False
        
    def toggle_filters():
        st.session_state.show_more_filters = not st.session_state.show_more_filters

    ### Faturamento
    faturamento_min = st.sidebar.number_input('Fat. Mín.:', value=0)
    faturamento_max = st.sidebar.number_input('Fat. Máx.:', value=df['FATURAMENTO'].max())

    if faturamento_min > faturamento_max:
        st.sidebar.error('O faturamento mínimo tem de ser menor que o faturamento máximo.')

    ### Data
    start_date = pd.to_datetime(st.sidebar.date_input('Data Inicial', value=df['DATA'].min()))
    end_date   = pd.to_datetime(st.sidebar.date_input('Data Final',   value=df['DATA'].max()))

    if start_date > end_date:
        st.sidebar.error('A data final tem que ser anterior à data inicial.')

    ### Categóricos
    filters = {
        'CLIENTE' : st.sidebar.multiselect(
         'Clientes:',
         options = df['CLIENTE'].unique(),
         default = []
         ),
        
        'COMERCIAL' : st.sidebar.multiselect(
         'Comercial:',
         options = df['COMERCIAL'].unique(),
         default = []
        ),
        
        'TIPO DE PRODUTO' : st.sidebar.multiselect(
         'Tipo de Produto:',
         options = df['TIPO DE PRODUTO'].unique(),
         default = []
        ),
        
        'SEGMENTO' : st.sidebar.multiselect(
         'Segmento:',
         options = df['SEGMENTO'].unique(),
         default = []
        ),
        
        'MERCADO' : st.sidebar.multiselect(
         'Mercado:',
         options = df['MERCADO'].unique(),
         default = []
        ),
        
        'ESTADO' : st.sidebar.multiselect(
         'Estado:',
         options = df['UF'].unique(),
         default = []
        ),
        
        'PAÍS' : st.sidebar.multiselect(
         'País:',
         options = df['PAÍS'].unique(),
         default = []
        )
    }

    ## Filtros Opcionais
    st.sidebar.button('Mais Filtros...', on_click=toggle_filters)
    if st.session_state.show_more_filters:
        st.sidebar.subheader('Filtros Adicionais')
        opt_filters = {
            'MUNICÍPIO' : st.sidebar.multiselect(
             'Município:',
             options = df['MUNICÍPIO'].unique(),
             default = []
            ),
            
            'CONTINENTE' : st.sidebar.multiselect(
             'Continente:',
             options = df['CONTINENTE'].unique(),
             default = []
            ),
            
            'ICMS' : st.sidebar.multiselect(
             'Dentro/Fora do Estado:',
             options = df['ICMS'].unique(),
             default = []
            ),
            
            'PRODUTO' : st.sidebar.multiselect(
             'Produto:',
             options = df['PRODUTO'].unique(),
             default = []
            ),
            
            'ORIGEM DO PRODUTO' : st.sidebar.multiselect(
             'Origem do Produto:',
             options = df['ORIGEM DO PRODUTO'].unique(),
             default = []
            ),
        }  

    ## Aplicar filtros
    df_filtered = df

    for column, selected_values in filters.items():
        if selected_values:
            df_filtered = df_filtered[df_filtered[column].isin(selected_values)]

    df_filtered = df_filtered[
        (df_filtered['DATA'] >= start_date) &
        (df_filtered['DATA'] <= end_date)
    ]

    df_filtered = df_filtered[
        (df_filtered['FATURAMENTO'] >= faturamento_min) &
        (df_filtered['FATURAMENTO'] <= faturamento_max)
    ]

    # Página Principal
    st.header('Estatísticas (com filtros)')

    ## Faturamento
    faturamento = int(df_filtered['FATURAMENTO'].sum())

    st.subheader(f'Faturamento: R$ {faturamento:,}')

    st.markdown('---')

    ## Curva ABC Clientes
    df_abc_cliente = df_filtered.groupby('CLIENTE').agg({
        'FATURAMENTO': 'sum',
        'ABC CLIENTE': 'first'
    }).sort_values(by='FATURAMENTO', ascending=False).reset_index()

    df_abc_cliente['FAT CUM'] = df_abc_cliente['FATURAMENTO'].cumsum()
    
    color_map = {'A': 'yellow', 'B': 'gray', 'C': 'orange'}
    colors = df_abc_cliente['ABC CLIENTE'].map(color_map)
    
    fig_curva_abc_cliente = go.Figure()
    
    fig_curva_abc_cliente.add_trace(go.Bar(
        x = df_abc_cliente['CLIENTE'],
        y = df_abc_cliente['FAT CUM'],
        marker_color = colors
    ))
    
    fig_curva_abc_cliente.update_layout(
        title = 'Curva ABC: Clientes',
        yaxis_title = 'Faturamento Cumulativo',
        xaxis_title = 'Cliente',
        xaxis = dict(showticklabels = False),
        width = 1000,
        height = 600,
    )
    
    st.plotly_chart(fig_curva_abc_cliente)
    
    ## Curva ABC Produtos
    df_abc_produto = df_filtered.groupby('PRODUTO').agg({
        'FATURAMENTO': 'sum',
        'ABC PRODUTO': 'first'
    }).sort_values(by='FATURAMENTO', ascending=False).reset_index()

    df_abc_produto['FAT CUM'] = df_abc_produto['FATURAMENTO'].cumsum()
    
    color_map = {'A': 'yellow', 'B': 'gray', 'C': 'orange'}
    colors = df_abc_produto['ABC PRODUTO'].map(color_map)
    
    fig_curva_abc_produto = go.Figure()
    
    fig_curva_abc_produto.add_trace(go.Bar(
        x = df_abc_produto['PRODUTO'],
        y = df_abc_produto['FAT CUM'],
        marker_color = colors
    ))
    
    fig_curva_abc_produto.update_layout(
        title = 'Curva ABC: Produtos',
        yaxis_title = 'Faturamento Cumulativo',
        xaxis_title = 'Produto',
        xaxis = dict(showticklabels = False),
        width = 1000,
        height = 600,
    )
    
    st.plotly_chart(fig_curva_abc_produto)

    ## Top 5 Clientes
    top5_clientes = (
        df_filtered.groupby('CLIENTE')['FATURAMENTO'].sum().sort_values(ascending = False)[:5]
    )

    fig_top5_clientes = px.bar(
        top5_clientes,
        x = top5_clientes.index,
        y = 'FATURAMENTO',
        title = '<b>Top 5 Clientes (Faturamento)</b>',
        template = 'plotly_white'
    )

    ## Top 5 Produtos
    top5_produtos = (
        df_filtered.groupby('PRODUTO')['QUANTIDADE'].sum().sort_values(ascending = False)[:5]
    )

    fig_top5_produtos = px.bar(
        top5_produtos,
        x = top5_produtos.index,
        y = 'QUANTIDADE',
        title = '<b>Top 5 Produtos (Quantidade)</b>',
        template = 'plotly_white'
    )

    left_col, right_col = st.columns(2)
    with left_col:
        st.plotly_chart(fig_top5_clientes)
    with right_col:
        st.plotly_chart(fig_top5_produtos)

    st.markdown('---')

    ## Faturamento por Mês
    faturamento_por_mes = (
        df_filtered.groupby('MÊS', sort=False)['FATURAMENTO'].sum()   
    )

    fig_faturamento_por_mes = px.bar(
        faturamento_por_mes,
        x = faturamento_por_mes.index,
        y = 'FATURAMENTO',
        title = '<b>Faturamento por Mês</b>',
        template = 'plotly_white'
    )

    ## Quantidade por Mês
    quantidade_por_mes = (
        df_filtered.groupby('MÊS', sort=False)['QUANTIDADE'].sum()   
    )

    fig_quantidade_por_mes = px.bar(
        quantidade_por_mes,
        x = quantidade_por_mes.index,
        y = 'QUANTIDADE',
        title = '<b>Quantidade por Mês</b>',
        template = 'plotly_white'
    )

    left_col, right_col = st.columns(2)
    with left_col:
        st.plotly_chart(fig_faturamento_por_mes)
    with right_col:
        st.plotly_chart(fig_quantidade_por_mes)

    st.markdown('---')

    ## Frações do Mercado
    fracao_mercado = (
        df_filtered.groupby('MERCADO')['FATURAMENTO'].sum()   
    )

    fig_fracao_mercado = px.pie(
        fracao_mercado,
        names = fracao_mercado.index,
        values = 'FATURAMENTO',
        title = '<b>Faturamento por Mercado</b>',
        template = 'plotly_white'
    )

    fracao_uf = (
        df_filtered.groupby('UF')['FATURAMENTO'].sum()   
    )

    fig_fracao_uf = px.pie(
        fracao_uf,
        names = fracao_uf.index,
        values = 'FATURAMENTO',
        title = '<b>Faturamento por UF</b>',
        template = 'plotly_white'
    )

    left_col, right_col = st.columns(2)
    with left_col:
        st.plotly_chart(fig_fracao_mercado)
    with right_col:
        st.plotly_chart(fig_fracao_uf)

with tabs[1]:

    st.header('Maiores Faturamentos')

    top30_faturamentos = df[['CLIENTE','FATURAMENTO']].sort_values('FATURAMENTO', ascending = False)
    top30_faturamentos = top30_faturamentos.reset_index(drop=True)
    st.dataframe(top30_faturamentos)

with tabs[2]:

    st.header('Tabela Completa:')

    st.dataframe(df_filtered)

with tabs[3]:
    
    st.header('Baixar Folha do Excel')
    
    file_name = st.text_input('Nome do Arquivo:', 'Banco de Dados PPack', key='Nome do Arquivo')
    
    if st.button('Converter para Excel...'):
        
        if file_name:
            
            # Add .xlsx
            file_name = file_name + '.xlsx'
            
            with st.spinner(text='Convertendo...'):
                
                # Save file
                excel_data = df_to_excel(df)
                st.download_button(
                    label='Baixar Excel',
                    data=excel_data,
                    file_name=file_name,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            
        else:
            st.error('Insira o nome do arquivo.')
            

with tabs[4]:
    
    st.header('Recomendações (EM CONSTRUÇÃO)')
    
    # Encode categorical
    df_rec = df.copy()
    label_encoders = {}
    for column in ['CLIENTE', 'PRODUTO', 'TIPO DE PRODUTO']:
        le = LabelEncoder()
        df_rec[column] = le.fit_transform(df_rec[column])
        label_encoders[column] = le

    # Create client-item interaction matrix
    user_item_matrix = df_rec.pivot_table(index='CLIENTE', columns='PRODUTO', values='FATURAMENTO', aggfunc='sum').fillna(0)
    
    # Cosine similarity matrix between each pair of products
    item_similarity = cosine_similarity(user_item_matrix.T)
    item_similarity_df = pd.DataFrame(item_similarity, index=user_item_matrix.columns, columns=user_item_matrix.columns) 
    
    # Generate recommendations
    def get_recommendations(client, user_item_matrix, item_similarity_df, top_n=5):
        
        # Get products client has already purchased
        client_data = user_item_matrix.loc[client_id]
        client_purchased = client_data[client_data > 0].index.tolist()	# list of products client has purchased
        
        # Calculate scores for products not purchased by client
        scores = {}
        for product in user_item_matrix.columns:
            if product not in client_purchased:
                product_score = sum(item_similarity_df[product][other_product] * client_data[other_product]
                                    for other_product in client_purchased)
                scores[product] = product_score
                
        # Sort products based on scores
        ranked_products = sorted(scores.items(), key=lambda x: x[1], reverse=True)
        return ranked_products[:top_n]
    
    clientes = df['CLIENTE'].unique()
    option = st.selectbox(
        'Cliente', clientes)
    
    client_id = label_encoders['CLIENTE'].transform([option])[0]
    recommendations = get_recommendations(client_id, user_item_matrix, item_similarity_df)

    # Corrected code to decode product IDs back to original product names
    recommended_products = pd.DataFrame([(label_encoders['PRODUTO'].inverse_transform([prod_id])[0], score) for prod_id, score in recommendations],
                                        columns = ['PRODUTO', 'PONTUAÇÃO']).
    
    st.write('Top Produtos Recomendados para ' + option)
    st.write(recommended_products)
#     i = 0
#     for product, score in recommended_products:
#         i += 1
#         st.write(str(i) + f'. Produto: {product}, Pontuação: {score}')
    
